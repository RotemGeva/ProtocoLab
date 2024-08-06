import logging
import os
import shutil
import subprocess
import glob
import openpyxl
import win32com.client
import sys
import tarfile
from pathlib import Path
import re
from tkinter import messagebox


def resource_path(relative_path):
    # """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


class Compare:
    def __init__(self, parameters):
        logging.info(f'***** Compare *****')
        try:
            self.parameters = parameters
            self.mr_name = Path(self.parameters.path_for_req).name.replace('_Requirements.xlsx', '')
            logging.info(f'MR: {self.mr_name}')
            self.extract_tar()
            self.move_relevant_protocols()
            if 'ProtocolExtractor' not in os.listdir('Data'):
                logging.info(f'Protocol Extractor tool not exist in {os.getcwd()}\\Data')
                self.prepare_protocol_extractor_for_compare()
            self.activate_protocol_extractor()
            self.move_protocol_extractor_output()
            self.prepare_for_comparison()
            self.compare()
            logging.info(f'***** Compare Done *****')
            sys.exit()
        except Exception as err:
            logging.error(f'Compare Error: {err}')
            sys.exit(1)
            #messagebox.showerror('ERROR', f'Compare Error: {err}')

    def extract_tar(self):
        logging.info(f'Start to extract TAR file')
        try:
            my_tar = tarfile.open(self.parameters.path_for_tar)
        except Exception as err:
            logging.info(f'Could not load TAR file.\nError: {err}')
            raise Exception(f'Could not load TAR file.\nError: {err}')
        try:
            my_tar.extractall(f'{os.getcwd()}\\Data\\temp\\Compare\\{self.mr_name}')  # specify which folder to extract to
        except Exception as err:
            logging.info(f'Could not extract TAR file to {os.getcwd()}\\Data\\temp\\Compare\\{self.mr_name}.\nError: {err}')
            raise Exception(f'Could not extract TAR file to {os.getcwd()}\\Data\\temp\\Compare\\{self.mr_name}.\nError: {err}')
        logging.info(f'TAR file extracted successfully to {os.getcwd()}\\Data\\temp\\Compare\\{self.mr_name}')
        my_tar.close()

    def move_relevant_protocols(self):
        logging.info(f'Start to move only relevant protocols to folder {os.path.dirname(self.parameters.path_for_req)}\\ForCompare')

        # Get a list of all sheets in the file and drove it into a variable
        try:
            requirement_file = openpyxl.load_workbook(self.parameters.path_for_req)
        except Exception as err:
            logging.info(f'Could not open Excel file.\nError: {err}')
            raise Exception(f'Could not open Excel file.\nError: {err}')
        try:
            list_of_all_sheets = requirement_file.sheetnames
        except Exception as err:
            logging.info(f'Could not receive sheets name from Excel file.\nError: {err}')
            raise Exception(f'Could not receive sheets name from Excel file.\nError: {err}')
        logging.info(f'Relevant protocols found from sheet names: {list_of_all_sheets}')

        # Get a list of all protocols in the file and drove it into a variable
        list_of_all_protocols = os.listdir(f'{os.getcwd()}\\Data\\temp\\Compare\\{self.mr_name}')
        logging.info(f'List of all extracted protocols in {os.getcwd()}\\Data\\temp\\Compare\\{self.mr_name}: {list_of_all_protocols}')

        if f'ForCompare' in os.listdir(f'{os.path.dirname(self.parameters.path_for_req)}'):
            logging.info(f'ForCompare folder already exist in {os.path.dirname(self.parameters.path_for_req)}')
            try:
                shutil.rmtree(f'{os.path.dirname(self.parameters.path_for_req)}\\ForCompare')
            except Exception as err:
                logging.info(f'Could not delete ForCompare folder from {os.path.dirname(self.parameters.path_for_req)}.\nError: {err}')
                raise Exception(f'Could not delete ForCompare folder from {os.path.dirname(self.parameters.path_for_req)}.\nError: {err}')
            logging.info(f'Existed ForCompare folder deleted')

        logging.info(f'Start moving relevant protocolos to: {os.path.dirname(self.parameters.path_for_req)}\\ForCompare')
        for sheet_name in list_of_all_sheets:
            for protocol_name in list_of_all_protocols:
                if sheet_name == re.match("^adult_other_(.+?)_\d+_\d+$", str(protocol_name)).group(1):
                    try:
                        shutil.copytree(f'{os.getcwd()}\\Data\\temp\\Compare\\{self.mr_name}\\{protocol_name}',
                                        f'{os.path.dirname(self.parameters.path_for_req)}\\ForCompare\\{protocol_name}')
                    except Exception as err:
                        logging.info(f'Could not move protocol folder {protocol_name} to {os.path.dirname(self.parameters.path_for_req)}\\ForCompare.\nError: {err}')
                        raise Exception(f'Could not move protocol folder {protocol_name} to {os.path.dirname(self.parameters.path_for_req)}\\ForCompare.\nError: {err}')
                    logging.info(f'Protocol {protocol_name} moved to {os.path.dirname(self.parameters.path_for_req)}\\ForCompare')
        logging.info(f'All relevant protocols for MR {self.mr_name} moved successfully')

    def prepare_protocol_extractor_for_compare(self):
        logging.info(f'Starting to move Protocol Extractor tool to {os.getcwd()}\\Data')
        try:
            temp_folder_list = os.listdir('\\\\192.100.100.116\\Public\\Testing\\Tools\\ProtocolExtractor\\bin\\V1.3')
            shutil.copytree(f'\\\\192.100.100.116\\Public\\Testing\\Tools\\ProtocolExtractor\\bin\\V1.3\\{temp_folder_list[len(temp_folder_list) - 1]}', f'Data\\ProtocolExtractor')
        except Exception as err:
            logging.info(f'Could not move Protocol Extractor tool to {os.getcwd()}\\Data.\nError: {err}')
            raise Exception(f'Could not move Protocol Extractor tool to {os.getcwd()}\\Data.\nError: {err}')
        logging.info(f'Protocol Extractor tool moved to {os.getcwd()}\\Data successfully')

    def activate_protocol_extractor(self):
        # remove all xls files from extractor folder
        logging.info(f'Removing all .xlsx files from Protocol Extractor folder')
        try:
            list_of_xls_files = glob.glob(f'{os.getcwd()}\\Data\\ProtocolExtractor\\*.xlsx')
            for xls_file in list_of_xls_files:
                os.remove(xls_file)
        except Exception as err:
            logging.info(f'failed to remove all .xlsx files from Protocol Extractor folder.\nError: {err}')
            raise Exception(f'failed to remove all .xlsx files from Protocol Extractor folder.\nError: {err}')
        logging.info(f'All .xlsx files removed from Protocol Extractor folder')
        logging.info(f'Running Protocol Extractor tool, commend: {os.getcwd()}\\Data\\ProtocolExtractor\\ProtocolExtractor.exe -f {os.getcwd()}\\Data\\{self.mr_name}\\ForCompare -m {os.getcwd()}\\Data\\ProtocolExtractor\\fields_map.ini')
        try:
            subprocess.run(f'{os.getcwd()}\\Data\\ProtocolExtractor\\ProtocolExtractor.exe -f {os.getcwd()}\\Data\\{self.mr_name}\\ForCompare -m {os.getcwd()}\\Data\\ProtocolExtractor\\fields_map.ini', cwd=f'{os.getcwd()}\\Data\\ProtocolExtractor')
        except Exception as err:
            logging.info(f'failed to run Protocol Extractor tool.\nError: {err}')
            raise Exception(f'failed to run Protocol Extractor tool.\nError: {err}')
        logging.info(f'Protocol Extractor tool done')

    def move_protocol_extractor_output(self):
        for file in os.listdir(f'{os.getcwd()}\\Data\\ProtocolExtractor'):
            if file.startswith('protocols_'):
                logging.info(f'Copying Protocol Extractor tool output')
                try:
                    shutil.copy2(f'{os.getcwd()}\\Data\\ProtocolExtractor\\{file}', f'{os.getcwd()}\\Data\\{self.mr_name}\\ForCompare')
                    logging.info(f'{file} moved from {os.getcwd()}\\Data\\ProtocolExtractor\\{file} to {os.getcwd()}\\Data\\{self.mr_name}\\ForCompare')
                except Exception as err:
                    logging.info(f'Failed to move Protocol Extractor tool output.\nError: {err}')
                    raise Exception(f'Failed to move Protocol Extractor tool output.\nError: {err}')
                try:
                    os.rename(f'{os.getcwd()}\\Data\\{self.mr_name}\\ForCompare\\{file}', f'{os.getcwd()}\\Data\\{self.mr_name}\\ForCompare\\{self.mr_name}_ForCompare.xlsx')
                    logging.info(f'{file} renamed to {self.mr_name}_ForCompare.xlsx')
                except Exception as err:
                    logging.info(f'Failed to rename Protocol Extractor tool output.\nError: {err}')
                    raise Exception(f'Failed to rename Protocol Extractor tool output.\nError: {err}')

    def prepare_for_comparison(self):
        logging.info(f'Preparing for comparison')
        logging.info(f'Create comparison file, copy {self.mr_name}_Requirements.xlsx and save as {self.mr_name}_Comparison.xlsx')
        try:
            shutil.copy2(f'{os.getcwd()}\\Data\\{self.mr_name}\\{self.mr_name}_Requirements.xlsx', f'{os.getcwd()}\\Data\\{self.mr_name}\\{self.mr_name}_Comparison.xlsx')
        except Exception as err:
            logging.info(f'Failed to copy {self.mr_name}_Requirements.xlsx file.\nError: {err}')
            raise Exception(f'Failed to copy {self.mr_name}_Requirements.xlsx file.\nError: {err}')
        logging.info(f'Open {self.mr_name}_ForCompare.xlsx using Python')
        try:
            for_compare_file = openpyxl.load_workbook(f'{os.getcwd()}\\Data\\{self.mr_name}\\ForCompare\\{self.mr_name}_ForCompare.xlsx')
        except Exception as err:
            logging.info(f'Failed to open {self.mr_name}_ForCompare.xlsx.\nError: {err}')
            raise Exception(f'Failed to open {self.mr_name}_ForCompare.xlsx.\nError: {err}')
        # Got a list of all sheets in the file and drove it into a variable
        sheets = for_compare_file.sheetnames
        # Deleting first sheet
        logging.info(f'Removing 1st sheet from {self.mr_name}_ForCompare.xlsx, sheet name: {for_compare_file[f"{sheets[0]}"]}')
        for_compare_file.remove(for_compare_file[f'{sheets[0]}'])
        # Saved file with changes (deleted page)
        for_compare_file.save(f'{os.getcwd()}\\Data\\{self.mr_name}\\ForCompare\\{self.mr_name}_ForCompare.xlsx')
        logging.info(f'Sheet removed and file {self.mr_name}_ForCompare.xlsx saved')

        # macros
        try:
            excel = win32com.client.dynamic.Dispatch('Excel.Application')
        except Exception as err:
            logging.info(f'Failed to open Excel software.\nError: {err}')
            raise Exception(f'Failed to open Excel software.\nError: {err}')
        # open macro excel file, macros will run from it
        logging.info(f'Open macro excel file using Python, file: {self.parameters.macro_excel_file}')
        try:
            excel.workbooks.Open(Filename=resource_path(f"{self.parameters.macro_excel_file}"))
        except Exception as err:
            logging.info(f'Failed to open macro excel file: {self.parameters.macro_excel_file}.\nError: {err}')
            raise Exception(f'Failed to open macro excel file: {self.parameters.macro_excel_file}.\nError: {err}')
        # open the file to run macros on
        logging.info(f'Open {self.mr_name}_ForCompare.xlsx for macros')
        try:
            for_compare_file = excel.workbooks.Open(Filename=f'{os.getcwd()}\\Data\\{self.mr_name}\\ForCompare\\{self.mr_name}_ForCompare.xlsx')
        except Exception as err:
            logging.info(f'Failed to open excel file: {self.mr_name}_ForCompare.\nError: {err}')
            raise Exception(f'Failed to open macro excel file: {self.mr_name}_ForCompare.xlsx.\nError: {err}')
        # run macro
        try:
            excel.Run(f"'{self.parameters.macro_excel_file}'!EqualCellSizeForAllSheets")
        except Exception as err:
            logging.info(f'Failed to run macro: EqualCellSizeForAllSheets.\nError: {err}')
            raise Exception(f'Failed to run macro: EqualCellSizeForAllSheets.\nError: {err}')
        try:
            excel.Run(f"'{self.parameters.macro_excel_file}'!CenterForAllSheets")
        except Exception as err:
            logging.info(f'Failed to run macro: CenterForAllSheets.\nError: {err}')
            raise Exception(f'Failed to run macro: CenterForAllSheets.\nError: {err}')
        # save the file after macros
        try:
            for_compare_file.Close(SaveChanges=1)
            logging.info(f'{self.mr_name}_ForCompare.xlsx Saved')
        except Exception as err:
            logging.info(f'Failed to save {self.mr_name}_ForCompare.xlsx.\nError: {err}')
            raise Exception(f'Failed to save {self.mr_name}_ForCompare.xlsx.\nError: {err}')
        excel.Quit()
        logging.info(f'All excel open files closed')

    def compare(self):
        # macros
        try:
            excel = win32com.client.dynamic.Dispatch('Excel.Application')
        except Exception as err:
            logging.info(f'Failed to open Excel software.\nError: {err}')
            raise Exception(f'Failed to open Excel software.\nError: {err}')
        # open macro excel file, macros will run from it
        logging.info(f'Open macro excel file using Python, file: {self.parameters.macro_excel_file}')
        try:
            excel.workbooks.Open(Filename=resource_path(f"{self.parameters.macro_excel_file}"))
        except Exception as err:
            logging.info(f'Failed to open macro excel file: {self.parameters.macro_excel_file}.\nError: {err}')
            raise Exception(f'Failed to open macro excel file: {self.parameters.macro_excel_file}.\nError: {err}')
        # open the file to run macros on
        logging.info(f'Open {self.mr_name}_Comparison.xlsx for macros')
        try:
            comparison_file = excel.workbooks.Open(Filename=f'{os.getcwd()}\\Data\\{self.mr_name}\\{self.mr_name}_Comparison.xlsx')
        except Exception as err:
            logging.info(f'Failed to open excel file: {self.mr_name}_Comparison.\nError: {err}')
            raise Exception(f'Failed to open macro excel file: {self.mr_name}_Comparison.xlsx.\nError: {err}')
        # run macro
        try:
            excel.Run(f"'{self.parameters.macro_excel_file}'!ReqToTestDocForAllSheets")
        except Exception as err:
            logging.info(f'Failed to run macro: ReqToTestDocForAllSheets.\nError: {err}')
            raise Exception(f'Failed to run macro: ReqToTestDocForAllSheets.\nError: {err}')
        try:
            excel.Run(f"'{self.parameters.macro_excel_file}'!PythonAutomaticCopyPasteToAllSheets", str(f'{os.getcwd()}\\Data\\{self.mr_name}\\ForCompare\\{self.mr_name}_ForCompare.xlsx'), str(f'{os.getcwd()}\\Data\\{self.mr_name}\\{self.mr_name}_Comparison.xlsx'))
        except Exception as err:
            logging.info(f'Failed to run macro: PythonAutomaticCopyPasteToAllSheets.\nError: {err}')
            raise Exception(f'Failed to run macro: PythonAutomaticCopyPasteToAllSheets.\nError: {err}')
        try:
            excel.Run(f"'{self.parameters.macro_excel_file}'!CompareForAllSheets")
        except Exception as err:
            logging.info(f'Failed to run macro: CompareForAllSheets.\nError: {err}')
            raise Exception(f'Failed to run macro: CompareForAllSheets.\nError: {err}')

        # save the file after macros
        try:
            comparison_file.Close(SaveChanges=1)
            logging.info(f'{self.mr_name}_Comparison.xlsx Saved')
        except Exception as err:
            logging.info(f'Failed to save {self.mr_name}_Comparison.xlsx.\nError: {err}')
            raise Exception(f'Failed to save {self.mr_name}_Comparison.xlsx.\nError: {err}')
        excel.Quit()
        logging.info(f'All excel open files closed')